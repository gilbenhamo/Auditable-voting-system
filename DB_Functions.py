import openpyxl
from token_generator import generateToken, getPayload, verifyToken
from enum import Enum
import random
import string
from cipher_funcs import *

PARTY1 = "Democrats"
PARTY2 = "Republicans"

VOTERS_DB = "votersDB.xlsx"
VOTES_DB = "votesDB.xlsx"
KEY = ""

def initKey(key):
    # Init the key
    global KEY
    KEY = key

def getSheet(name):
    # Load workbook and return it
    book = openpyxl.load_workbook(name)
    return book.active,book

def searchInSheet(sheet,value):
    # Search value in A colunm on sheet
    for i in range(1,sheet.max_row+1):
        if value == sheet['A'+str(i)].value:
            return i
    return -1

def isVoted(sheet,index):
    # Check if this index is exists
    return sheet['B'+str(index)].value

def vote(sheet, id):
    # Check if voter can vote
    index=searchInSheet(sheet,id)
    if index==-1:
        return "The id wasnt found"

    if isVoted(sheet,index):
        return "The preson already voted"
        
    return index

def writeVoter(choos, poll, IV, bookVotes):
    # Create vote and write it on DB

    # Create payload
    payload = {
        'choos' : choos,
        'poll' : poll,
        'IV' : IV
    }
    key = IV + KEY
    token = generateToken(payload, key)

    # Encrypt data
    ciphertext = encrypt(choos,key)
    hex_ciphertext = ciphertext.hex()
    row = [token, hex_ciphertext]

    # Write it on DB
    bookVotes.active.append(row)
    bookVotes.save(VOTES_DB)
    return token

def updateVotersDB(index):
    # Change the index in the B column to TRUE
    voters_sheet,voters_book = getSheet(VOTERS_DB)
    voters_sheet['B'+str(index)] = True
    voters_book.save(VOTERS_DB)

def generateIV():
    # Generate string with length 5
    return ''.join(random.choices(string.ascii_lowercase, k=5))


def getAllVotes():
    # Calculate resaults
    sheet,votes_book = getSheet(VOTES_DB)
    count1 = count2 = count3 = 0
    amountOfVotes = sheet.max_row
    for i in range(1,amountOfVotes+1):
        token = sheet['A'+str(i)].value
        IV = getPayload(token)['IV']

        if not verifyToken(token,KEY):
            count3+=1
        
        else:
            try:
                choos = decrypt(sheet['B'+str(i)].value ,IV+KEY)
                if PARTY1 == choos:
                    count1+=1
                elif PARTY2 == choos:
                    count2+=1
                else:
                    count3+=1
            except Exception:
                count3+=1

    return count1/amountOfVotes*100, count2/amountOfVotes*100, count3/amountOfVotes*100

def clearDB(DB_name):
    # Clear DB
    sheet,book = getSheet(DB_name)
    sheet.delete_rows(1,sheet.max_row)
    book.save(DB_name)

def initVotersToFalse():
    # Make all voters didnt vote
    sheet,book = getSheet(VOTERS_DB)
    for i in range(1,sheet.max_row+1):
        sheet['B'+str(i)] = False
    book.save(VOTERS_DB)
    
initVotersToFalse()